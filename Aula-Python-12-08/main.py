from openpyxl import Workbook, load_workbook
import random #https://pypi.org/project/randon/

def configs():  

     try:
          
       wb = load_workbook(filename = 'campo.xlsx')
       config = wb['config']

     except:
          
        wb = Workbook()

        config = wb.create_sheet('config')
        config.cell(column = 1, row = 1, value = 'linhas')
        config.cell(column = 2, row = 1, value = 5)
        config.cell(column = 1, row = 2, value = 'colunas')
        config.cell(column = 2, row = 2, value = 5)
        config.cell(column = 1, row = 3, value = 'dificuldade')
        config.cell(column = 2, row = 3, value = 2)

     linhas = config.cell(column = 2, row = 1).value
     colunas = config.cell(column = 2, row = 2).value
     dificuldade = config.cell(column = 2, row = 3).value

     wb.save('campo.xlsx')
     
     config = {
        
        'linhas' : int(linhas),
        'colunas' : int(colunas),
        'dificuldade' : int(dificuldade)
     }
     
     return config

def CriarTabuleiro():
    config = configs()
    quantLinhas = config['linhas']
    quantColunas = config['colunas']
    bombas = CalcularBomba()
    tabuleiro = []
    cont = 1

   # 0 -> casas vazias | -1 -> bombas
    for i in range(0, int(quantLinhas)):
      linha = []
      for j in range(0, int(quantColunas)):
         print('{:^3}'.format('[🏡]'), end='')
         if cont in bombas:
            linha.append('-1') 
         else:
            linha.append('0')

         cont += 1
      print()
      tabuleiro.append(linha)

    gravarTabuleiro(tabuleiro)

def gravarTabuleiro(tabuleiro):
   try:
      wb = load_workbook(filename = "campo.xlsx")
   except:
      wb = Workbook()

   try:
      abajogo = wb['Jogo']
   except:
      abajogo = wb.create_sheet('Jogo')
   print()
   abajogo.delete_rows(1, abajogo.max_row)
   abajogo.delete_cols(1, abajogo.max_column)

   for linha in range(1, len(tabuleiro) + 1):
      for coluna in range(1, len(tabuleiro[0]) + 1):
         abajogo.cell(row=linha, column=coluna, value=tabuleiro[linha - 1][coluna -1])

   wb.save('campo.xlsx')

def ConfigCampo():

   NovaQuantColunas = input('\n Quantas colunas: ')
   NovaQuantLinhas = input(' Quantas linhas: ')
   NovaDificuldade = input('\n 1- Facil \n 2- Medio \n 3- Dificil \n\n Escolha a dificuldade: ')

   try:

      wb = load_workbook(filename = 'campo.xlsx')
      config = wb['config']

   except:

      wb = Workbook()
      config = wb.create_sheet('config')
      
   config.cell(column = 1, row = 1, value = 'linhas')
   config.cell(column = 2, row = 1, value = NovaQuantLinhas)
   config.cell(column = 1, row = 2, value = 'colunas')
   config.cell(column = 2, row = 2, value = NovaQuantColunas)
   config.cell(column = 1, row = 3, value = 'dificuldade')
   config.cell(column = 2, row = 3, value = NovaDificuldade)


   wb.save('campo.xlsx')

def CalcularBomba():
   '''
   Descubra o valor de 'casas' no tabuleiro e a partir desse total calcule a quantidade de bombas que vai ter no tabuleiro
   Fácil -> 15%, Médio -> 30 %, Dificil-> 50% de bombas
   '''

   config = configs()
   totalCasas = config['linhas'] * config['colunas']

   if config['dificuldade'] == 1:   
      quantBomba = totalCasas * 0.15    
   elif config['dificuldade'] == 2:     
      quantBomba = totalCasas * 0.3
   elif config['dificuldade'] == 3:   
      quantBomba = totalCasas * 0.5

   # print ('\nTotal de Bombas: {}' .format(int(quantBomba)))

   #Bombas Aleatórias [1,2,3],[4,5,6],[7,8,9]

   bombas = []
   while True:
      posicao = random.randint(1, totalCasas)
      if posicao not in bombas:
         bombas.append(posicao)
      if len(bombas) == int(quantBomba):
         break

   return bombas
      
def Jogar():
   CriarTabuleiro()

   wb = load_workbook(filename= 'campo.xlsx')
   jogo = wb['Jogo']
   maxLinha = jogo.max_row
   maxColuna = jogo.max_column
   gameOver = False
   jgPossivel = (maxLinha * maxColuna) - len(CalcularBomba())
   historico = 0

   while True:

      linhaJogada = int(input("\nEscolha a Linha: "))
      colunaJogada = int(input("\nEscolha a Coluna: "))

      jogada = int(jogo.cell(row=linhaJogada, column=colunaJogada).value)
      if jogada == 0:
         jogo.cell(row=linhaJogada, column=colunaJogada, value= 1)
         historico += 1
      elif jogada == -1:
         gameOver = True
      elif jogada == 1:
         print()
         print('Jogada Já Efetuada, Tente Novamente')

      wb.save('campo.xlsx')
      
      # 🏡, ✅, 💥
      for linha in range(1, maxLinha + 1):
         for coluna in range( 1, maxColuna + 1):
            casa = jogo.cell(row=linha, column=coluna).value
            if int(casa) == 0 or (int(casa) == -1 and gameOver == False):
               print('{:^3}'.format('[🏡]'), end='')
            elif int(casa) == 1:
               print('{:^3}'.format('[✅]'), end='')
            elif int(casa) <= -1 and gameOver == True:
               print('{:^3}'.format('[💥]'), end='')
         print()

      if gameOver == True:
         print()
         print('Você Perdeu ❌')
         print()
         break
      if historico == jgPossivel:
         print()
         print("Parabéns, Você Ganhou!! 🥇")
         print()
         break
   if historico == jgPossivel or gameOver == True:
      op = input("Você deseja jogar novamente ? (s/n): ")
      if op == 's' or 'S':
         Jogar()

while True:

#Crie o menu de acesso do jogo com as seguintes opções "jogar,Configurar, Sair"

   opcao = input ('\n 1- jogar \n 2- Configurar \n 3- Sair \n\n Escolha uma das opções: ')

   if opcao == '1':
      Jogar()

   elif opcao == '2':
      print('\n Configurar')
      ConfigCampo()

   elif opcao == '3':
      print('\n Sair')
      break

   else:
      print('\033[31mErro!!\033[0m')