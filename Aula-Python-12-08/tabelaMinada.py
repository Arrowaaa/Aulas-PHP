from openpyxl import Workbook, load_workbook

def configs():  

     try:
          
       wb = load_workbook(filename = 'campo.xlsx')
       config = wb['config']

     except:
          
        wb = Workbook()

        config = wb.create_sheet('config')
        config.cell(column = 1, row = 1, value = 'linhas')
        config.cell(column = 2, row = 1, value = 5)
        config.cell(column = 1, row = 2, value = 'colunas')
        config.cell(column = 2, row = 2, value = 5)
        config.cell(column = 1, row = 3, value = 'dificuldade')
        config.cell(column = 2, row = 3, value = 2)

     linhas = config.cell(column = 2, row = 1).value
     colunas = config.cell(column = 2, row = 2).value
     dificuldade = config.cell(column = 2, row = 3).value

     wb.save('campo.xlsx')
     
     config = {
        
        'linhas' : int(linhas),
        'colunas' : int(colunas),
        'dificuldade' : int(dificuldade)
     }
     
     return config

def CriarTabuleiro(quantLinhas,quantColunas):
    
    for i in range(0, int(quantLinhas)):
      for j in range(0, int(quantColunas)):
         print('{:5}'.format(' [ ] '), end = '')
      print()

def ConfigCampo():

   NovaQuantColunas = input('\n Quantas colunas: ')
   NovaQuantLinhas = input(' Quantas linhas: ')
   NovaDificuldade = input('\n 1- Facil \n 2- Medio \n 3- Dificil \n\n Escolha a dificuldade: ')

   try:

      wb = load_workbook(filename = 'campo.xlsx')
      config = wb['config']

   except:

      wb = Workbook()
      config = wb.create_sheet('config')
      
   config.cell(column = 1, row = 1, value = 'linhas')
   config.cell(column = 2, row = 1, value = NovaQuantLinhas)
   config.cell(column = 1, row = 2, value = 'colunas')
   config.cell(column = 2, row = 2, value = NovaQuantColunas)
   config.cell(column = 1, row = 3, value = 'dificuldade')
   config.cell(column = 2, row = 3, value = NovaDificuldade)


   wb.save('campo.xlsx')

def CalcularBomba():
   
   
   totalLinhas, totalColunas = configs()
   totalCasas = config['linhas'] * config['colunas']

   # if (dificuldade == '1'):   
   #    quantCasa = quantLinha * quantColuna
   #    quantTotalBomba = quantCasa * 0.15
      
   # elif (dificuldade == '2'):     
   #    quantCasa = quantLinha * quantColuna
   #    quantTotalBomba = quantCasa * 0.3
      
   # elif (dificuldade == '3'):   
   #    quantCasa = quantLinha * quantColuna
   #    quantTotalBomba = quantCasa * 0.5
      
config = configs()
CriarTabuleiro(config['linhas'], config['colunas'])

while True:

   opcao = input ('\n 1- jogar \n 2- Configurar \n 3- Sair \n\n Escolha: ')

   if opcao == '1':
      print('\n jogar')

   elif opcao == '2':
      print('\n Configurar')
      ConfigCampo()

      config = configs()
      CriarTabuleiro(config['linhas'], config['colunas'])

   elif opcao == '3':
      print('\n Sair')
      break

   else:
      print('Deu erro sla porque')